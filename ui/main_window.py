from __future__ import annotations
import os
from datetime import date
from PyQt6 import QtWidgets, QtCore, QtGui
from pathlib import Path
import pandas as pd
from typing import TYPE_CHECKING, Optional, Callable

from app_frantoio.resources import resource_path
from app_frantoio.models.moliture_model import MolitureModel
from app_frantoio.util.time_utils import fmt_ts
from app_frantoio.ui.reports_widget import ReportsPage

if TYPE_CHECKING:
    from app_frantoio.core.repository import HybridRepository


# ---------------------- Icona app ----------------------
def _find_icon_path() -> Path:
    try:
        p = Path(resource_path("app.ico"))
        if p.exists():
            return p
    except Exception:
        pass
    base = Path(QtCore.QCoreApplication.applicationDirPath())
    for c in [
        base / "resources" / "app.ico",
        base / "app_frantoio" / "resources" / "app.ico",
        base / "app.ico",
    ]:
        if c.exists():
            return c
    return Path()


def _app_icon() -> QtGui.QIcon:
    p = _find_icon_path()
    return QtGui.QIcon(str(p)) if p.exists() else QtGui.QIcon()


# ---------------------- LoginDialog ----------------------
class LoginDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowIcon(_app_icon())
        self.setWindowTitle("Accesso")
        self.setModal(True)
        layout = QtWidgets.QVBoxLayout(self)

        form = QtWidgets.QFormLayout()
        self.email = QtWidgets.QLineEdit()
        self.email.setPlaceholderText("email@example.com")
        self.passw = QtWidgets.QLineEdit()
        self.passw.setEchoMode(QtWidgets.QLineEdit.EchoMode.Password)
        self.passw.setPlaceholderText("Password (min 6 caratteri)")
        form.addRow("Email", self.email)
        form.addRow("Password", self.passw)
        layout.addLayout(form)

        self.status = QtWidgets.QLabel("")
        layout.addWidget(self.status)

        btns = QtWidgets.QHBoxLayout()
        self.btnCancel = QtWidgets.QPushButton("Annulla")
        self.btnLogin = QtWidgets.QPushButton("Accedi")
        self.btnLogin.setDefault(True)
        self.btnLogin.setFocus()
        btns.addStretch(1)
        btns.addWidget(self.btnCancel)
        btns.addWidget(self.btnLogin)
        layout.addLayout(btns)

        self.btnLogin.clicked.connect(self.accept)
        self.btnCancel.clicked.connect(self.reject)

    def get_credentials(self):
        return self.email.text().strip(), self.passw.text().strip()


# ---------------------- Worker generico (thread pool) ----------------------
class _WorkerSignals(QtCore.QObject):
    done = QtCore.pyqtSignal(object)   # result
    error = QtCore.pyqtSignal(str)     # error message


class _FuncWorker(QtCore.QRunnable):
    def __init__(self, fn: Callable, *args, **kwargs):
        super().__init__()
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = _WorkerSignals()
        self.setAutoDelete(True)

    @QtCore.pyqtSlot()
    def run(self):
        try:
            res = self.fn(*self.args, **self.kwargs)
            self.signals.done.emit(res)
        except Exception as e:
            self.signals.error.emit(str(e))


# ---------------------- Delegate Abbuono ----------------------
class AbbuonoDelegate(QtWidgets.QStyledItemDelegate):
    editingStarted = QtCore.pyqtSignal()

    def createEditor(self, parent, option, index):
        self.editingStarted.emit()  # pausa timer subito
        edit = QtWidgets.QLineEdit(parent)
        edit.setAlignment(QtCore.Qt.AlignmentFlag.AlignRight)
        rx = QtCore.QRegularExpression(r"^\s*\d{0,12}([.,]\d{0,2})?\s*$")
        edit.setValidator(QtGui.QRegularExpressionValidator(rx, edit))
        edit.returnPressed.connect(lambda: self.commitData.emit(edit))
        edit.returnPressed.connect(lambda: self.closeEditor.emit(
            edit, QtWidgets.QAbstractItemDelegate.EndEditHint.NoHint))
        edit.editingFinished.connect(lambda: self.commitData.emit(edit))
        edit.editingFinished.connect(lambda: self.closeEditor.emit(
            edit, QtWidgets.QAbstractItemDelegate.EndEditHint.NoHint))
        return edit

    def setEditorData(self, editor, index):
        val = index.model().data(index, QtCore.Qt.ItemDataRole.EditRole)
        try:
            editor.setText(f"{float(val or 0.0):.2f}")
        except Exception:
            editor.setText("0.00")

    def setModelData(self, editor, model, index):
        txt = editor.text().strip().replace(",", ".")
        try:
            value = float(txt) if txt else 0.0
        except ValueError:
            value = 0.0
        model.setData(index, value, QtCore.Qt.ItemDataRole.EditRole)


# ---------------------- MainWindow ----------------------
class MainWindow(QtWidgets.QMainWindow):
    def __init__(self, cfg, repo: HybridRepository):
        super().__init__()
        self.cfg = cfg
        self.repo = repo
        self.euro_per_kg = float(cfg.get("euro_per_kg", 0.30))
        self.poll_ms = int(cfg.get("poll_ms", 3000))
        self._export_path = None
        self._busy = False  # fix cursore/ri-entrance

        # tracking selezione per evitare "rimbalzi" dopo refresh
        self._last_selected_id: Optional[str] = None
        self._sel_serial: int = 0  # aumenta ad ogni cambio selezione

        self.setWindowTitle("Gestione Moliture - PC")
        self.resize(1200, 700)
        self.setWindowIcon(_app_icon())

        # TABS
        self.tabs = QtWidgets.QTabWidget(self)

        # --- TAB 1: Lista ---
        page_list = QtWidgets.QWidget()
        vbox = QtWidgets.QVBoxLayout(page_list)

        controls = QtWidgets.QHBoxLayout()
        controls.addWidget(QtWidgets.QLabel("Data:"))
        self.date_edit = QtWidgets.QDateEdit(calendarPopup=True)
        self.date_edit.setDate(QtCore.QDate.currentDate())
        self.date_edit.setDisplayFormat("dd/MM/yyyy")
        controls.addWidget(self.date_edit)
        controls.addStretch(1)
        self.btn_refresh = QtWidgets.QPushButton("Aggiorna")
        self.btn_export = QtWidgets.QPushButton("Export Excel")
        controls.addWidget(self.btn_refresh)
        controls.addWidget(self.btn_export)
        vbox.addLayout(controls)

        self.table = QtWidgets.QTableView()
        self.model = MolitureModel([], self.euro_per_kg, self)
        self.table.setModel(self.model)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.SingleSelection)
        self.table.horizontalHeader().setStretchLastSection(True)
        vbox.addWidget(self.table)

        self.table.setEditTriggers(
            QtWidgets.QAbstractItemView.EditTrigger.DoubleClicked
            | QtWidgets.QAbstractItemView.EditTrigger.SelectedClicked
            | QtWidgets.QAbstractItemView.EditTrigger.EditKeyPressed
        )

        self.abbuono_delegate = AbbuonoDelegate(self.table)
        self.table.setItemDelegateForColumn(3, self.abbuono_delegate)

        self.model.dataChanged.connect(self.on_model_changed)

        self.abbuono_delegate.editingStarted.connect(self._on_editor_start)
        self.abbuono_delegate.commitData.connect(self._on_editor_commit)
        self.abbuono_delegate.closeEditor.connect(self._on_editor_closed)

        # traccia ogni cambio selezione dell'utente
        self.table.selectionModel().selectionChanged.connect(self._on_selection_changed)

        footer = QtWidgets.QHBoxLayout()
        self.lbl_tot = QtWidgets.QLabel("Totale kg: 0.00   |   Abbuoni: € 0.00")
        footer.addWidget(self.lbl_tot)
        footer.addStretch(1)
        footer.addWidget(QtWidgets.QLabel("Pagamento rapido:"))
        self.cmb_pagamento_quick = QtWidgets.QComboBox()
        self.cmb_pagamento_quick.addItems(["", "Contanti", "POS", "Assegno", "Olio"])
        self.btn_set_pagamento = QtWidgets.QPushButton("Imposta su riga selezionata")
        self.btn_delete_row = QtWidgets.QPushButton("Cancella record selezionato")
        footer.addWidget(self.cmb_pagamento_quick)
        footer.addWidget(self.btn_set_pagamento)
        footer.addWidget(self.btn_delete_row)
        vbox.addLayout(footer)

        self.tabs.addTab(page_list, "Lista")

        # --- TAB 2: Report ---
        self.reports_page = ReportsPage(self.repo, self.euro_per_kg, self)
        self.tabs.addTab(self.reports_page, "Report")

        self.setCentralWidget(self.tabs)

        # Thread pool
        self._pool = QtCore.QThreadPool.globalInstance()

        # Timer refresh tabella (auto): NON mostra cursore
        self.timer = QtCore.QTimer(self)
        self.timer.setInterval(self.poll_ms)
        self.timer.timeout.connect(lambda: self.refresh_data(show_cursor=False))

        # Timer mirror (senza cleanup)
        self.sync_timer = QtCore.QTimer(self)
        minutes = max(1, int(self.cfg.get("mirror_interval_minutes", 5)))
        self.sync_timer.setInterval(minutes * 60 * 1000)
        self.sync_timer.timeout.connect(self.auto_sync)

        # Signals
        self.date_edit.dateChanged.connect(self.refresh_data)          # show_cursor=True (default)
        self.btn_refresh.clicked.connect(self.refresh_data)            # show_cursor=True (default)
        self.btn_export.clicked.connect(self.export_excel)
        self.btn_set_pagamento.clicked.connect(self.on_set_pagamento_clicked)
        self.btn_delete_row.clicked.connect(self.on_delete_clicked)

        # Avvio
        QtCore.QTimer.singleShot(200, self.refresh_data)              # primo load (show_cursor=True)
        QtCore.QTimer.singleShot(1000, self.auto_sync)
        self.timer.start()
        self.sync_timer.start()

    # ---------------------- Helpers selezione ----------------------
    def _on_selection_changed(self, selected: QtCore.QItemSelection, deselected: QtCore.QItemSelection):
        cur = self._current_selected_id()
        if cur:
            self._last_selected_id = cur
        self._sel_serial += 1

    def _current_selected_id(self) -> Optional[str]:
        sel = self.table.selectionModel().selectedRows()
        if not sel:
            return None
        row = self.model.row_at(sel[0].row())
        return row.get("id")

    def _reselect_by_id(self, rid: Optional[str]):
        if not rid:
            return
        sm = self.table.selectionModel()
        if sm is None:
            return
        rows = self.model.all_rows()
        for r_idx, r in enumerate(rows):
            if r.get("id") == rid:
                idx = self.model.index(r_idx, 0)
                sm.select(
                    idx,
                    QtCore.QItemSelectionModel.SelectionFlag.ClearAndSelect
                    | QtCore.QItemSelectionModel.SelectionFlag.Rows,
                )
                self.table.setCurrentIndex(idx)
                self.table.scrollTo(idx, QtWidgets.QAbstractItemView.ScrollHint.PositionAtCenter)
                break

    def _selected_date(self) -> date:
        qd = self.date_edit.date()
        return date(qd.year(), qd.month(), qd.day())

    # ---------------------- Busy UI / Cursor fix ----------------------
    def _set_busy(self, busy: bool):
        if busy == getattr(self, "_busy", False):
            return  # nessun cambio di stato

        self._busy = busy

        widgets = [
            self.btn_refresh, self.btn_export,
            self.btn_set_pagamento, self.btn_delete_row,
            self.date_edit, self.table
        ]
        for w in widgets:
            w.setEnabled(not busy)

        if busy:
            if self.timer.isActive():
                self.timer.stop()
            QtGui.QGuiApplication.setOverrideCursor(QtCore.Qt.CursorShape.WaitCursor)
        else:
            # svuota TUTTI gli override residui
            try:
                while QtGui.QGuiApplication.overrideCursor() is not None:
                    QtGui.QGuiApplication.restoreOverrideCursor()
            except Exception:
                pass
            if not self.timer.isActive():
                self.timer.start()

    # ---------------------- Refresh (in background) ----------------------
    def refresh_data(self, show_cursor: bool = True):
        # non refreshare durante l'editing o se già occupati
        if self.table.state() == QtWidgets.QAbstractItemView.State.EditingState:
            return
        if self._busy:
            return

        d = self._selected_date()
        prev_id = self._current_selected_id()
        vscroll = self.table.verticalScrollBar().value()

        # seriale selezione all'avvio del refresh
        sel_serial_start = self._sel_serial

        if show_cursor:
            self._set_busy(True)
        else:
            # niente icona: solo flag per evitare re-ingressi
            self._busy = True

        def _fetch():
            return self.repo.fetch_day(d)

        worker = _FuncWorker(_fetch)
        worker.signals.done.connect(
            lambda rows: self._on_refresh_done(rows, prev_id, vscroll, show_cursor, sel_serial_start)
        )
        worker.signals.error.connect(lambda msg: self._on_refresh_error(msg, show_cursor))
        self._pool.start(worker)

    @QtCore.pyqtSlot(object)
    def _on_refresh_done(self, rows, prev_id, vscroll, show_cursor: bool, sel_serial_start: int):
        try:
            rows_list = list(rows or [])
            try:
                rows_list.sort(key=lambda r: int(float(r.get("dataOra") or 0)))
            except Exception:
                pass
            self.model.set_rows(rows_list)
            self.update_total(rows_list)
            self.table.verticalScrollBar().setValue(vscroll)

            # Se l'utente ha cambiato selezione durante il fetch, rispetta la NUOVA scelta
            if self._sel_serial != sel_serial_start and self._last_selected_id:
                self._reselect_by_id(self._last_selected_id)
            else:
                # altrimenti ripristina la selezione precedente (se esiste)
                self._reselect_by_id(prev_id)
        finally:
            if show_cursor:
                self._set_busy(False)
            else:
                self._busy = False

    @QtCore.pyqtSlot(str)
    def _on_refresh_error(self, msg: str, show_cursor: bool):
        if show_cursor:
            self._set_busy(False)
        else:
            self._busy = False
        QtWidgets.QMessageBox.warning(self, "Lettura dati", f"Errore: {msg}")

    def update_total(self, rows):
        total_kg = 0.0
        for r in rows:
            try:
                total_kg += float(r.get("weight") or r.get("peso") or 0)
            except Exception:
                pass
        self.lbl_tot.setText(f"Totale kg: {total_kg:.2f}")

    # ---------------------- Auto-sync (mirror in background) ----------------------
    def auto_sync(self):
        if self._busy:  # non accavallare con refresh
            return

        def _mirror():
            return self.repo.mirror_only()

        worker = _FuncWorker(_mirror)
        worker.signals.done.connect(lambda mirrored: self.statusBar().showMessage(
            f"Sync eseguito: salvati {mirrored} record su SQLite.", 5000))
        worker.signals.error.connect(lambda msg: self.statusBar().showMessage(f"Sync fallito: {msg}", 5000))
        self._pool.start(worker)

    # ---------------------- Azioni UI ----------------------
    def on_set_pagamento_clicked(self):
        self.timer.stop()
        try:
            idxs = self.table.selectionModel().selectedRows()
            if not idxs:
                QtWidgets.QMessageBox.information(self, "Pagamento", "Seleziona una riga prima.")
                return
            idx = idxs[0]
            row = self.model.row_at(idx.row())
            pagamento = self.cmb_pagamento_quick.currentText()
            rid = row.get("id")
            if not rid:
                raise ValueError("ID mancante")
            source = row.get("_source", "firebase")
            self.repo.update_pagamento(rid, pagamento, source)
            row["pagamento"] = pagamento
            self.model.dataChanged.emit(
                self.model.index(idx.row(), 5),
                self.model.index(idx.row(), 5),
                [QtCore.Qt.ItemDataRole.DisplayRole],
            )
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Errore aggiornamento", f"{e}")
        finally:
            self.timer.start()

    def on_delete_clicked(self):
        idxs = self.table.selectionModel().selectedRows()
        if not idxs:
            QtWidgets.QMessageBox.information(self, "Cancella", "Seleziona una riga prima.")
            return

        idx = idxs[0]
        row = self.model.row_at(idx.row())
        rid = row.get("id")
        if not rid:
            QtWidgets.QMessageBox.warning(self, "Cancella", "ID record mancante.")
            return

        name = row.get("name") or row.get("nome") or ""
        w = row.get("weight") or row.get("peso") or 0
        try:
            w = float(w)
        except Exception:
            w = 0.0
        when = fmt_ts(row.get("dataOra"))
        source = row.get("_source", "firebase")

        msg = (
            "Vuoi cancellare il record selezionato dalla sua sorgente?\n\n"
            f"Nome: {name}\nPeso: {w:.2f} kg\nOra: {when}\n"
            f"Sorgente: {source.upper()}"
        )
        reply = QtWidgets.QMessageBox.question(
            self,
            "Conferma cancellazione",
            msg,
            QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
        )
        if reply != QtWidgets.QMessageBox.StandardButton.Yes:
            return

        try:
            self.repo.delete_record(rid, source)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Errore cancellazione", str(e))
            return

        self.refresh_data(show_cursor=False)  # refresh “silenzioso” dopo cancellazione
        self.statusBar().showMessage("Record cancellato.", 3000)

    def on_model_changed(self, topLeft: QtCore.QModelIndex, bottomRight: QtCore.QModelIndex, roles):
        row = topLeft.row()
        col = topLeft.column()
        r = self.model.row_at(row)
        rid = r.get("id")
        if not rid:
            return
        source = r.get("_source", "firebase")

        if col == 3:  # Abbuono
            try:
                val = float(r.get("abbuono") or 0.0)
            except Exception:
                val = 0.0
            QtCore.QTimer.singleShot(0, lambda rid=rid, val=val, src=source: self._save_abbuono(rid, val, src))

        elif col == 5:  # Pagamento
            val = r.get("pagamento") or ""
            QtCore.QTimer.singleShot(0, lambda rid=rid, val=val, src=source: self._save_pagamento(rid, val, src))

    # --- gestione editor per evitare refresh in edit ---
    def _on_editor_start(self, *args):
        self.timer.stop()

    def _on_editor_commit(self, *args):
        self.timer.stop()

    def _on_editor_closed(self, *args):
        if not self.timer.isActive():
            self.timer.start()

    # --- salvataggi posticipati ---
    def _save_abbuono(self, rid: str, val: float, source: str):
        self.timer.stop()
        try:
            if source == "firebase":
                self.repo.update_abbuono(rid, val, "firebase")
                try:
                    self.repo.update_abbuono(rid, val, "sqlite")
                except Exception:
                    pass
            else:
                self.repo.update_abbuono(rid, val, "sqlite")

            self.statusBar().showMessage(f"Abbuono aggiornato a € {val:.2f}", 2000)
            QtCore.QTimer.singleShot(200, lambda: self.refresh_data(show_cursor=False))
        except Exception as e:
            self.statusBar().showMessage(f"Errore salvataggio abbuono: {e}", 4000)
        finally:
            if not self.timer.isActive():
                self.timer.start()

    def _save_pagamento(self, rid: str, val: str, source: str):
        self.timer.stop()
        try:
            if source == "firebase":
                self.repo.update_pagamento(rid, val, "firebase")
                try:
                    self.repo.update_pagamento(rid, val, "sqlite")
                except Exception:
                    pass
            else:
                self.repo.update_pagamento(rid, val, "sqlite")

            self.statusBar().showMessage(f"Pagamento aggiornato: {val}", 2000)
            QtCore.QTimer.singleShot(200, lambda: self.refresh_data(show_cursor=False))
        except Exception as e:
            self.statusBar().showMessage(f"Errore salvataggio pagamento: {e}", 4000)
        finally:
            if not self.timer.isActive():
                self.timer.start()

    # ---------------------- Export Excel ----------------------
    def export_excel(self):
        rows = self.model.all_rows()
        if not rows:
            QtWidgets.QMessageBox.information(self, "Export", "Nessun dato da esportare.")
            return

        d = self._selected_date()
        sheet_name = d.strftime("%d.%m.%Y")

        df = pd.DataFrame(rows)
        if "name" not in df.columns and "nome" in df.columns:
            df["name"] = df["nome"]
        if "weight" not in df.columns and "peso" in df.columns:
            df["weight"] = df["peso"]

        df["Ora"] = df.get("dataOra", "").apply(fmt_ts)
        df["abbuono"] = pd.to_numeric(df.get("abbuono", 0.0), errors="coerce").fillna(0.0)
        df["Prezzo (€)"] = pd.to_numeric(df["weight"], errors="coerce").fillna(0.0) * float(self.euro_per_kg)

        cols = ["name", "weight", "Prezzo (€)", "abbuono", "Ora", "pagamento"]
        cols = [c for c in cols if c in df.columns]
        df = df[cols]
        df.columns = ["Nome", "Peso (Kg)", "Prezzo (€)", "Abbuono (€)", "Ora", "Metodo di Pagamento"]

        if not self._export_path:
            path, _ = QtWidgets.QFileDialog.getSaveFileName(
                self, "Seleziona file Excel", "molitura.xlsx", "Excel (*.xlsx)"
            )
            if not path:
                return
            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"
            self._export_path = path
        file_path = self._export_path

        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment

            if os.path.exists(file_path):
                with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df.to_excel(writer, index=False, sheet_name=sheet_name)
            else:
                with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name=sheet_name)

            wb = load_workbook(file_path)
            ws = wb[sheet_name]

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[col_letter]:
                    text = "" if cell.value is None else str(cell.value)
                    if len(text) > max_len:
                        max_len = len(text)
                ws.column_dimensions[col_letter].width = max_len + 2

            wb.save(file_path)
            QtWidgets.QMessageBox.information(
                self, "Export", f"Esportato nel foglio '{sheet_name}' di {os.path.basename(file_path)}"
            )
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Errore export", str(e))
